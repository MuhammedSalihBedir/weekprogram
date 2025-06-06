import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from weasyprint import HTML, CSS
from pdf2image import convert_from_bytes

def combine_sequenced_lectures(week):
    for day in week:
        for hour in week[day]:
            for l, lec in enumerate(week[day][hour]):
                week[day][hour][l]["count"] = 1


    for day in week:
        reversed_hours = list(week[day].keys())[::-1]
        for h, hour in enumerate(reversed_hours[:-1]):
            for l in range(len(week[day][hour])-1, -1, -1):
                lec = week[day][hour][l]
                for l2, lec2 in enumerate(week[day][reversed_hours[h+1]]):
                    id_check = lec["id"] == lec2["id"]
                    prof_check = lec["professor"]["id"] == lec2["professor"]["id"]
                    classroom_check = lec["classroom"]["id"] == lec2["classroom"]["id"]
                    locked_check = ("locked" in lec and "locked" in lec2) or ("locked" not in lec and "locked" not in lec2)
                    if id_check and prof_check and classroom_check and locked_check:
                        week[day][reversed_hours[h+1]][l2]["count"] += week[day][hour][l]["count"]
                        week[day][hour].pop(l)
                        break
    
    return week

def tableize_combined_week_by_year(combined_week, remove_unnecessary_nones = False):
    cols_availibility_counter = {}
    table = {
        "rows": {},
        "cols": {},
    }
    global_row_pointer = 0

    for day in combined_week:
        table["rows"][day] = {}

        for hour in combined_week[day]:
            table["rows"][day][hour] = {}

            if not remove_unnecessary_nones:
                for year in cols_availibility_counter:
                    for i, item in enumerate(cols_availibility_counter[year]):
                        if item != 0:
                            table["cols"][year][i].append(None)

            lectures = combined_week[day][hour][::-1]

            for l in range(len(lectures)-1, -1, -1):
                lec = lectures[l]

                if lec["year"] not in cols_availibility_counter:
                    cols_availibility_counter[lec["year"]] = []
                    table["cols"][lec["year"]] = []

                for year in cols_availibility_counter:

                    if lec["year"] != year:
                        continue

                    for i, item in enumerate(cols_availibility_counter[year]):
                        if item == 0:
                            table["cols"][year][i].append(lec)
                            cols_availibility_counter[year][i] = lec["count"]
                            lectures.pop(l)
                            break
                    else:
                        table["cols"][year].append([{} for _ in range(global_row_pointer)] + [lec])
                        cols_availibility_counter[year].append(lec["count"])


            for year in cols_availibility_counter:
                for i, item in enumerate(cols_availibility_counter[year]):
                    if item == 0:
                        table["cols"][year][i].append({})
                    else:
                        cols_availibility_counter[year][i] -= 1

            global_row_pointer += 1

    table["cols"] = dict(sorted(table["cols"].items())) # to sort keys

    return table

def build_week_html_content(tableized_week, years, hours):
    with open("static/week-schedual-table.css", "r") as f:
        style = "<style>" + f.read() + "</style>"

    table_content = style + "<table>"

    header_content = """<tr>
    <th class="bordered_cell header_cell">Day</th>
    <th class="bordered_cell header_cell">Time</th>""" + "".join(
        f"""
        <th class="bordered_cell header_cell year_header" colspan="{len(tableized_week["cols"][year]) if year in tableized_week["cols"] else 1}">Year {year}</th>
        """
        for year in years # tableized_week["cols"]
    ) + """
</tr>"""

    global_table_row_pointer = 0

    for day in tableized_week["rows"]:
        table_content += header_content
        
        # days_content = f"""<td class="bordered_cell" rowspan="{len(tableized_week["rows"][day])+1}">{day}</td>"""
        days_content = f"""<td class="bordered_cell" rowspan="{len(hours)+1}">{day}</td>"""

        week_hours = list(tableized_week['rows'][day].keys())

        rows_content = []
        # for hour in tableized_week["rows"][day]:
        for hour in hours:
            if hour.is_rest:
                rows_content.append(f"""<tr><td class="bordered_cell">{hour.name}:00 ~ {hour.name}:50</td><td class="rest_hour" colspan="{sum(
                    (
                        1 if year not in tableized_week["cols"] else len(tableized_week["cols"][year])
                    )
                    for year in years
                )}">Break Hour</td>""")
            else:
                rows_content.append(f"""<tr><td class="bordered_cell">{hour.name}:00 ~ {hour.name}:50</td>""")
        
        # for year in tableized_week["cols"]:
        for year in years:
            if year not in tableized_week["cols"]:
                for l in range(len(rows_content)):
                # for l in range(len(hours)):
                    if hours[l].is_rest:
                        continue
                    rows_content[l] += f"<td day='{day}' hour='{week_hours[l]}' year='{year}' onclick='placeLecture(event)'></td>"
                continue

            for col in tableized_week["cols"][year]:
                for l, lec in enumerate(col[global_table_row_pointer : global_table_row_pointer + len(rows_content)]):
                    if hours[l].is_rest:
                        continue

                # for l, lec in enumerate(col[global_table_row_pointer : global_table_row_pointer + len(work_hours)]):
                    if lec is None:
                        continue

                    if lec == {}:
                        rows_content[l] += f"<td day='{day}' hour='{week_hours[l]}' year='{year}' onclick='placeLecture(event)'></td>"
                        continue

                    span = lec["count"]

                    rows_content[l] += f"""<td rowspan="{span}" day='{day}' hour='{week_hours[l]}' year='{year}' onclick='placeLecture(event)' class="bordered_cell lec_cell {("locked" in lec and lec["locked"]) * "lec_cell_locked"}">{lec["name"]}<br>{lec["professor"]["name"]}<br>{lec["classroom"]["name"]}</td>"""
                    # rows_content[l] += f"""<td rowspan="{span}" class="bordered_cell lec_cell">{lec["name"]}<br>{lec["professorName"]}</td>"""
        
        global_table_row_pointer += len(rows_content)

        for i in range(len(rows_content)):
            rows_content[i] += "</tr>"
        
        days_content += "".join(rows_content)

        table_content += days_content

    table_content += "</table>"

    return table_content

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max_length

def auto_adjust_row_height(ws):
    for row in ws.iter_rows():
        max_height = 15  # Default row height
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1  # Count line breaks
                max_height = max(max_height, lines * 15)  # Adjust for multiline text
        
        ws.row_dimensions[row[0].row].height = max_height

def get_max_dimentions(tableized_week, years):
    rows_num = 0
    for day in tableized_week["rows"]:
        rows_num += len(tableized_week["rows"][day]) + 1 # fore headers

    cols_num = 2
    for year in years: # tableized_week["cols"]:
        years_columns_num = len(tableized_week["cols"][year] )if year in tableized_week["cols"] else 1
        cols_num += years_columns_num
    
    return rows_num, cols_num

def build_week_excel_file(tableized_week, years, hours):
    header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    cell_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    break_fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
    alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # Center alignment
    day_alignment = Alignment(textRotation=90, horizontal="center", vertical="center")  # Center alignment
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    def add_header(ws, row, col):
        cell = ws.cell(row=1+row, column=1+col, value=f"Day")
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

        cell = ws.cell(row=1+row, column=1+col+1, value=f"Time")
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

        col += 2

        for year in years: # tableized_week["cols"]:
            years_columns_num = len(tableized_week["cols"][year] )if year in tableized_week["cols"] else 1
            
            cell = ws.cell(row=1+row, column=1+col, value=f"Year {year}")
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

            ws.merge_cells(
                start_row=1+row,
                start_column=1+col,
                end_row=1+row,
                end_column=1+col+years_columns_num -1
            )

            col += years_columns_num

    global_row_pointer = 0
    global_table_row_pointer = 0
    global_col_pointer = 0


    for day in tableized_week["rows"]:
        add_header(ws, global_row_pointer, global_col_pointer)
        global_row_pointer += 1

        cell = ws.cell(row=1+global_row_pointer, column=1+global_col_pointer, value=day)
        # cell.fill = cell_fill
        cell.alignment = day_alignment
        cell.border = border

        ws.merge_cells(
            start_row=1+global_row_pointer,
            start_column=1+global_col_pointer,
            end_row=1+global_row_pointer+len(tableized_week["rows"][day]) -1,
            end_column=1+global_col_pointer
        )

        local_row_pointer = global_row_pointer
        for hour in tableized_week["rows"][day]:
            cell = ws.cell(row=1+local_row_pointer, column=1+global_col_pointer+1, value=f"{hour}:00 ~ {hour}:50")
            # cell.fill = cell_fill
            cell.alignment = alignment
            cell.border = border
            local_row_pointer += 1
        
        hours_num = len(tableized_week["rows"][day])

        local_col_pointer = global_col_pointer
        for year in years: # tableized_week["cols"]:
            if year not in tableized_week["cols"]:
                local_col_pointer += 1
                continue

            for col in tableized_week["cols"][year]:
                local_row_pointer = global_row_pointer
                for i, lec in enumerate(col[global_table_row_pointer : global_table_row_pointer + hours_num]):
                    local_row_pointer += 1

                    if hours[i].is_rest:
                        cell = ws.cell(row=local_row_pointer, column=1+local_col_pointer+2, value="Break Hour")
                        cell.fill = break_fill
                        cell.alignment = alignment
                        cell.border = border
                        continue

                    if lec is None:
                        continue

                    if lec == {}:
                        continue

                    to_skip = lec["count"] -1
                    
                    cell = ws.cell(
                        row=local_row_pointer,
                        column=1+local_col_pointer+2,
                        # value=lec["name"] + "\n" + lec["professor"]["name"] + "\n" + lec["classroom"]["name"]
                        value=lec["name"] + "\n" + lec["professor"]["name"] + "\n" + str(lec["classroom"]["name"])
                        # value=lec["name"] + "\n" + lec["professorName"]
                    )
                    cell.fill = cell_fill
                    cell.alignment = alignment
                    cell.border = border
                    ws.merge_cells(
                        start_row=local_row_pointer,
                        start_column=1+local_col_pointer+2,
                        end_row=local_row_pointer+to_skip,
                        end_column=1+local_col_pointer+2
                    )

                    # local_row_pointer += to_skip
                
                local_col_pointer += 1
        
        global_row_pointer += hours_num
        global_table_row_pointer += hours_num

    rows_num, cols_num = get_max_dimentions(tableized_week, years)

    for i in range(rows_num):
        cell = ws.cell(row=1+i, column=cols_num)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, right=Side(style="thin"), bottom=b.bottom)

    for i in range(cols_num):
        cell = ws.cell(row=rows_num, column=1+i)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, right=b.right, bottom=Side(style="thin"))

    # auto_adjust_column_width(ws)
    # auto_adjust_row_height(ws)

    # wb.save("merged_cells.xlsx")
    return wb

def build_time_table_html_content(times_list, days, hours):
    html_string = "<table id='schedule-table'>"

    html_string += "<thead><tr>"
    for item in ["Time"] + days:
        html_string += f"<th value='{item}'>{item}</th>"
    html_string += "</tr></thead>"

    for item in hours:
        html_string += "<tr>"
        html_string += f"<td value='{item.name}'>{item.name}:00 ~ {item.name}:50</td>"

        if item.is_rest:
            html_string += f"""<td class="rest_hour" colspan="{len(days)}">Break Hour</td>"""
        else:
            for day in days:
                for time in times_list:
                    if time["day"] == day and time["hour"] == item.name:
                        html_string += f"<td class='selected' onclick='this.classList.toggle(\"selected\");'></td>"
                        break
                else:
                    html_string += f"<td onclick='this.classList.toggle(\"selected\");'></td>"
            
        html_string += "</tr>"

    foot_content = "<tfoot><tr><td><button onclick='saveFreeTime()'>Save</button></td>"

    for d in range(len(days)):
        foot_content += f"""<td>
    <button onclick="eraseDay({d+1})">Erase</button>
    <button onclick="toggleDay({d+1})">Toggle</button>
</td>"""
    
    foot_content += "</tr></tfoot>"

    html_string += foot_content
    html_string += "</table>"

    with open("static/timeprof-table.css", "r") as f:
        style = "<style>" + f.read() + "</style>"

    return style + html_string


def get_fitted_document_from_html(html_string, margin_mm=10):
    # 1) First pass on a virtually infinite canvas
    doc      = HTML(string=html_string)
    css_huge = CSS(string=f'''
        @page {{ size: 10000mm 10000mm; margin: {margin_mm}mm; }}
        body {{ margin: 0; }}
    ''')
    layout   = doc.render(stylesheets=[css_huge])
    page     = layout.pages[0]
    page_box = page._page_box

    w_px    = page_box.children[0].children[0].children[0].margin_width()
    h_px    = page_box.children[0].children[0].margin_height()

    # 3) Convert CSS px → mm (96 dpi → 25.4 mm/inch)
    mm_per_px = 25.4 / 96.0
    w_mm = w_px * mm_per_px + margin_mm * 2
    h_mm = h_px * mm_per_px + margin_mm * 2

    # 4) Second pass: tight crop + margin
    css_tight = CSS(string=f'''
        @page {{ size: {w_mm}mm {h_mm}mm; margin: {margin_mm}mm; }}
        body {{ margin: 0; }}
    ''')
    return doc, css_tight

def get_pdf_from_html(html_string, margin_mm=10):
    doc, css_tight = get_fitted_document_from_html(html_string, margin_mm)
    return doc.write_pdf(stylesheets=[css_tight], zoom=1)

def get_png_from_html(html_string, margin_mm=10):
    pdf_bytes = get_pdf_from_html(html_string, margin_mm)
    return convert_from_bytes(pdf_bytes)[0]
